import openpyxl

# loading the sheet and reading the content in the python file
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Task 1 : reading the supplier name and storing them in the dictionary

# dictionary name for this task1
product_per_supplier = {}
total_value_per_supplier = {}
product_inv_less_10 = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row ,1).value
    inventory_file = product_list.cell(product_row , 5)

    # calculation number of products per supplier
    if supplier_name in product_per_supplier:
        current_num_products = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name] = current_num_products + 1
    else:
        product_per_supplier[supplier_name] = 1

    # calculation total value of inventory per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price

    # logic products with inventory less than 10
    if inventory < 10:
        product_inv_less_10[product_num] = inventory

    # to add another cell in the file for the total
    inventory_file.value = inventory * price

print(product_per_supplier)
print(total_value_per_supplier)
print(product_inv_less_10)

inv_file.save("inventory_data_added.xlsx")


